import openpyxl
import glob
import os

xlsx_files = glob.glob(r"c:\Users\dogip\Downloads\AI-MOCK-INTERVIEW-main\AI-MOCK-INTERVIEW-main\e2e-tests\*.xlsx")
for f in xlsx_files:
    print(f"--- File: {os.path.basename(f)} ---")
    wb = openpyxl.load_workbook(f)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"Sheet: {sheetname}")
        # Print top 15 rows
        for r in range(1, min(16, sheet.max_row + 1)):
            row_vals = [sheet.cell(r, c).value for c in range(1, min(10, sheet.max_column + 1))]
            print(f"Row {r:02d}: {row_vals}")
        # Let's count status values (if there's a status column)
        # We can scan the row values to find where tests start and count PASSED / FAILED
        statuses = {}
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                val = str(sheet.cell(r, c).value)
                if val in ["PASSED", "FAILED", "SUCCESS", "FAIL"]:
                    statuses[val] = statuses.get(val, 0) + 1
        print(f"Status counts: {statuses}")
