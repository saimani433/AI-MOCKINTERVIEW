import openpyxl
import glob
import os

print("Listing excel files in e2e-tests:")
xlsx_files = glob.glob(r"c:\Users\dogip\Downloads\AI-MOCK-INTERVIEW-main\AI-MOCK-INTERVIEW-main\e2e-tests\*.xlsx")
for f in xlsx_files:
    print(f"File: {os.path.basename(f)}")
    wb = openpyxl.load_workbook(f)
    print(f"  Sheets: {wb.sheetnames}")
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"    Sheet '{sheetname}': {sheet.max_row} rows, {sheet.max_column} columns")
