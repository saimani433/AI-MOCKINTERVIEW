import os
import sys
import pytest
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class ExcelReporterPlugin:
    """
    A custom pytest plugin that intercepts test execution results 
    and logs them into a list for Excel spreadsheet compilation.
    """
    def __init__(self):
        self.results = []

    def pytest_runtest_logreport(self, report):
        # We only log result details during the 'call' phase (not setup or teardown)
        if report.when == "call":
            nodeid = report.nodeid
            test_name = nodeid.split("::")[-1]
            
            # Categorize the test cases based on function name
            category = "Unit / API Validation"
            if "ui_viewport" in test_name:
                category = "UI/UX Layout"
            elif "interview_creation" in test_name:
                category = "Functional Logic"
            elif "registration_validation" in test_name:
                category = "Boundary Validation"

            self.results.append({
                "Test ID": f"TC-{len(self.results) + 1:03d}",
                "Category": category,
                "Test Name": test_name,
                "Status": "PASSED" if report.passed else "FAILED",
                "Duration (s)": round(report.duration, 5),
                "Details": str(report.longrepr)[:200] if report.failed else "Check executed successfully."
            })

def run_suite_and_generate_excel():
    print("Starting execution of 400+ comprehensive test cases...")
    
    # Instantiate our custom pytest collector plugin
    reporter = ExcelReporterPlugin()
    
    # Run pytest programmatically on test_scenarios.py
    pytest.main(["-q", "test_scenarios.py"], plugins=[reporter])
    
    total_tests = len(reporter.results)
    print(f"Test suite finished. Captured {total_tests} test cases.")
    
    # Create a new workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Status Summary"
    
    # Enable gridlines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # -------------------------------------------------------------
    # Styling Definitions
    # -------------------------------------------------------------
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    
    title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep navy
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Medium navy
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
    
    status_font_pass = Font(name="Segoe UI", size=10, bold=True, color="375623")
    status_font_fail = Font(name="Segoe UI", size=10, bold=True, color="C00000")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # -------------------------------------------------------------
    # Sheet Title Banner
    # -------------------------------------------------------------
    ws.merge_cells("A1:F2")
    title_cell = ws["A1"]
    title_cell.value = "VocaVision AI Test Execution Report"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    
    # Set heights for title rows
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    
    # -------------------------------------------------------------
    # Summary Dashboard Block
    # -------------------------------------------------------------
    passed_count = sum(1 for r in reporter.results if r["Status"] == "PASSED")
    failed_count = total_tests - passed_count
    pass_rate = f"{(passed_count / total_tests) * 100:.1f}%" if total_tests > 0 else "0%"
    
    ws["A4"] = "Execution Summary"
    ws["A4"].font = Font(name="Segoe UI", size=11, bold=True, color="1F497D")
    
    ws["A5"] = "Total Test Cases"
    ws["B5"] = total_tests
    ws["A6"] = "Success (PASSED)"
    ws["B6"] = passed_count
    ws["A7"] = "Failures (FAILED)"
    ws["B7"] = failed_count
    ws["A8"] = "Pass Rate"
    ws["B8"] = pass_rate
    
    # Style summary labels
    for row in range(5, 9):
        ws.cell(row=row, column=1).font = Font(name="Segoe UI", size=10, bold=True)
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=2).font = Font(name="Segoe UI", size=10)
        ws.cell(row=row, column=2).alignment = center_align
        
    # Apply soft color fills to summary results
    ws.cell(row=6, column=2).font = status_font_pass
    ws.cell(row=6, column=2).fill = pass_fill
    ws.cell(row=7, column=2).font = status_font_fail
    ws.cell(row=7, column=2).fill = fail_fill
    ws.cell(row=8, column=2).font = Font(name="Segoe UI", size=10, bold=True, color="1F497D")
    
    # Empty spacer row
    ws.row_dimensions[9].height = 15
    
    # -------------------------------------------------------------
    # Column Headers (Moved to Row 10)
    # -------------------------------------------------------------
    headers = ["Test ID", "Category", "Test Name", "Status", "Duration (s)", "Details"]
    header_row = 10
    ws.row_dimensions[header_row].height = 25
    
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # -------------------------------------------------------------
    # Populate Test Results (Moved to Row 11 onwards)
    # -------------------------------------------------------------
    start_data_row = 11
    for idx, res in enumerate(reporter.results):
        current_row = start_data_row + idx
        ws.row_dimensions[current_row].height = 20
        
        # Test ID
        cell_id = ws.cell(row=current_row, column=1, value=res["Test ID"])
        cell_id.font = data_font
        cell_id.alignment = center_align
        
        # Category
        cell_cat = ws.cell(row=current_row, column=2, value=res["Category"])
        cell_cat.font = data_font
        cell_cat.alignment = left_align
        
        # Test Name
        cell_name = ws.cell(row=current_row, column=3, value=res["Test Name"])
        cell_name.font = data_font
        cell_name.alignment = left_align
        
        # Status
        cell_status = ws.cell(row=current_row, column=4, value=res["Status"])
        cell_status.alignment = center_align
        if res["Status"] == "PASSED":
            cell_status.font = status_font_pass
            cell_status.fill = pass_fill
        else:
            cell_status.font = status_font_fail
            cell_status.fill = fail_fill
            
        # Duration
        cell_dur = ws.cell(row=current_row, column=5, value=res["Duration (s)"])
        cell_dur.font = data_font
        cell_dur.alignment = center_align
        
        # Details
        cell_det = ws.cell(row=current_row, column=6, value=res["Details"])
        cell_det.font = data_font
        cell_det.alignment = left_align

    # -------------------------------------------------------------
    # Auto-Fit Columns Width
    # -------------------------------------------------------------
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid title merged cells length calculation skewing width
            if cell.row in [1, 2]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save the Workbook
    output_filename = "selenium_report.xlsx"
    ws.views.sheetView[0].showGridLines = True
    wb.save(output_filename)
    print(f"Excel report successfully generated: {os.path.abspath(output_filename)}")
    
    # Also save to test_execution_report.xlsx
    wb.save("test_execution_report.xlsx")
    print(f"Excel report successfully generated: {os.path.abspath('test_execution_report.xlsx')}")

if __name__ == "__main__":
    run_suite_and_generate_excel()
